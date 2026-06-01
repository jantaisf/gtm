from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers)
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill as PF

wb = Workbook()

# ── helpers ──────────────────────────────────────────────────────────────────
NAVY_HEX   = "1E3A5F"
WHITE_HEX  = "FFFFFF"
YELLOW_HEX = "FFFF00"
BLUE_HEX   = "0000FF"
BLACK_HEX  = "000000"
GREEN_HEX  = "008000"

ARIAL = "Arial"

FMT_CURRENCY  = '$#,##0;($#,##0);"-"'
FMT_CURRENCY0 = '$#,##0;($#,##0);"-"'
FMT_PCT       = '0.0%;(0.0%);"-"'
FMT_MULT      = '0.00x'

RED_LIGHT_HEX   = "FFC7CE"   # light red bg
GREEN_LIGHT_HEX = "C6EFCE"   # light green bg

def navy_title(ws, row, col, text, merge_end_col=8):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name=ARIAL, bold=True, size=12, color=WHITE_HEX)
    cell.fill = PatternFill("solid", fgColor=NAVY_HEX)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22
    if merge_end_col > col:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=merge_end_col)

def section_header(ws, row, col, text, merge_end_col=8):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name=ARIAL, bold=True, size=10, color=WHITE_HEX)
    cell.fill = PatternFill("solid", fgColor="2E4A7A")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 18
    if merge_end_col > col:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=merge_end_col)

def col_header(ws, row, col, text, wrap=True):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name=ARIAL, bold=True, size=9, color=BLACK_HEX)
    cell.fill = PatternFill("solid", fgColor="D9E1F2")
    cell.alignment = Alignment(horizontal="center", vertical="center",
                               wrap_text=wrap)
    ws.row_dimensions[row].height = 30

def label_cell(ws, row, col, text):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name=ARIAL, size=9, color=BLACK_HEX)
    cell.alignment = Alignment(horizontal="left", vertical="center")

def input_cell(ws, row, col, value, fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name=ARIAL, size=9, color=BLUE_HEX)
    cell.fill = PatternFill("solid", fgColor=YELLOW_HEX)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    if fmt:
        cell.number_format = fmt
    return cell

def formula_cell(ws, row, col, formula, fmt=None, color=BLACK_HEX):
    cell = ws.cell(row=row, column=col, value=formula)
    cell.font = Font(name=ARIAL, size=9, color=color)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    if fmt:
        cell.number_format = fmt
    return cell

def link_cell(ws, row, col, formula, fmt=None):
    cell = ws.cell(row=row, column=col, value=formula)
    cell.font = Font(name=ARIAL, size=9, color=GREEN_HEX)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    if fmt:
        cell.number_format = fmt
    return cell

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_border_range(ws, min_row, max_row, min_col, max_col):
    for r in range(min_row, max_row+1):
        for c in range(min_col, max_col+1):
            ws.cell(row=r, column=c).border = thin_border()

def set_col_widths(ws, widths):
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

# ════════════════════════════════════════════════════════════════════════════
# TAB 1 — INPUTS
# ════════════════════════════════════════════════════════════════════════════
ws_inp = wb.active
ws_inp.title = "Inputs"

navy_title(ws_inp, 1, 1, "Prisma Cloud cACV Comp Model — Inputs", merge_end_col=4)

# ── DEAL PARAMETERS ──────────────────────────────────────────────────────────
section_header(ws_inp, 3, 1, "DEAL PARAMETERS", merge_end_col=4)

col_header(ws_inp, 4, 1, "Parameter", wrap=False)
col_header(ws_inp, 4, 2, "Value")
col_header(ws_inp, 4, 3, "Notes", wrap=False)

deal_params = [
    ("ACV ($/yr)",                  300000,  FMT_CURRENCY,  "Annual Contract Value"),
    ("Contract term (years)",       3,        "0",           "Valid values: 1, 2, 3, 5"),
    ("Year 1 consumption rate",     0.80,    FMT_PCT,       "% of ACV consumed in Year 1"),
    ("Year 2 consumption rate",     0.85,    FMT_PCT,       "% of ACV consumed in Year 2"),
    ("Year 3 consumption rate",     0.90,    FMT_PCT,       "% of ACV consumed in Year 3"),
    ("Year 4 consumption rate",     0.90,    FMT_PCT,       "Only relevant for 5-yr deals"),
    ("Year 5 consumption rate",     0.90,    FMT_PCT,       "Only relevant for 5-yr deals"),
]

# Row map for Inputs (1-indexed)
# B5 = ACV, B6 = term, B7-B11 = consumption rates Yr1-5
DEAL_START_ROW = 5
for i, (lbl, val, fmt, note) in enumerate(deal_params):
    r = DEAL_START_ROW + i
    label_cell(ws_inp, r, 1, lbl)
    input_cell(ws_inp, r, 2, val, fmt)
    c = ws_inp.cell(row=r, column=3, value=note)
    c.font = Font(name=ARIAL, size=8, color="595959", italic=True)

# ── QUOTA STRUCTURE ───────────────────────────────────────────────────────────
section_header(ws_inp, 13, 1, "QUOTA STRUCTURE", merge_end_col=4)
col_header(ws_inp, 14, 1, "Parameter", wrap=False)
col_header(ws_inp, 14, 2, "Value")

quota_params = [
    ("AE total quota ($)",   1000000, FMT_CURRENCY),
    ("AE bookings weight",   0.70,    FMT_PCT),
    ("AE cACV weight",       0.30,    FMT_PCT),
    ("AM total quota ($)",   800000,  FMT_CURRENCY),
    ("AM bookings weight",   0.30,    FMT_PCT),
    ("AM cACV weight",       0.70,    FMT_PCT),
]

QUOTA_START_ROW = 15
# Row map: B15=AE quota, B16=AE bk wt, B17=AE cACV wt, B18=AM quota, B19=AM bk wt, B20=AM cACV wt
for i, (lbl, val, fmt) in enumerate(quota_params):
    r = QUOTA_START_ROW + i
    label_cell(ws_inp, r, 1, lbl)
    input_cell(ws_inp, r, 2, val, fmt)

# ── TERM MULTIPLIERS ─────────────────────────────────────────────────────────
section_header(ws_inp, 22, 1, "TERM MULTIPLIERS", merge_end_col=4)
col_header(ws_inp, 23, 1, "Term", wrap=False)
col_header(ws_inp, 23, 2, "Multiplier")

mults = [
    ("1-year multiplier", 1.00),
    ("2-year multiplier", 1.20),
    ("3-year multiplier", 1.35),
    ("5-year multiplier", 1.50),
]

MULT_START_ROW = 24
# B24=1yr, B25=2yr, B26=3yr, B27=5yr
for i, (lbl, val) in enumerate(mults):
    r = MULT_START_ROW + i
    label_cell(ws_inp, r, 1, lbl)
    input_cell(ws_inp, r, 2, val, "0.00x")

# ── FOOTER NOTE ───────────────────────────────────────────────────────────────
ws_inp.merge_cells(start_row=29, start_column=1, end_row=30, end_column=4)
note_cell = ws_inp.cell(row=29, column=1,
    value='Source: Arora (PANW Q2 FY2024 earnings): avg contract ~3 years; '
          'platformization deals 3–5 years. '
          'cACV = MIN(ACV × consumption_rate, ACV). '
          'expansion_signal_acv = MAX(ACV × rate − ACV, 0).')
note_cell.font = Font(name=ARIAL, size=8, italic=True, color="595959")
note_cell.alignment = Alignment(wrap_text=True, vertical="top")
ws_inp.row_dimensions[29].height = 30

set_col_widths(ws_inp, [28, 14, 38, 5])
apply_border_range(ws_inp, 4, 11, 1, 3)
apply_border_range(ws_inp, 14, 20, 1, 2)
apply_border_range(ws_inp, 23, 27, 1, 2)


# ════════════════════════════════════════════════════════════════════════════
# TAB 2 — AE MODEL
# ════════════════════════════════════════════════════════════════════════════
ws_ae = wb.create_sheet("AE Model")
navy_title(ws_ae, 1, 1, "Prisma Cloud cACV Comp Model — AE Model", merge_end_col=7)

# ── Cell references into Inputs (shorthand) ──
# Inputs!B5  = ACV
# Inputs!B6  = term
# Inputs!B7  = Yr1 rate, B8=Yr2, B9=Yr3, B10=Yr4, B11=Yr5
# Inputs!B15 = AE quota, B16=AE bk wt, B17=AE cACV wt
# Inputs!B18 = AM quota, B19=AM bk wt, B20=AM cACV wt
# Inputs!B24 = mult 1yr, B25=2yr, B26=3yr, B27=5yr

# Helper: get multiplier for a given term_years value (1,2,3,5) using CHOOSE or IFS
# We'll use IFS formula: =IFS(Inputs!B6=1,Inputs!B24, B6=2,..., B6=3,..., B6=5,...)
MULT_FORMULA = ("=IFS(Inputs!B6=1,Inputs!B24,"
                "Inputs!B6=2,Inputs!B25,"
                "Inputs!B6=3,Inputs!B26,"
                "Inputs!B6=5,Inputs!B27)")

# ── SECTION 1: Deal Economics by Term ────────────────────────────────────────
section_header(ws_ae, 3, 1, "SECTION 1 — Deal Economics by Term", merge_end_col=7)

hdrs = ["Term", "ACV ($/yr)", "TCV ($)",
        "Quota Credit — No Multiplier ($)",
        "Quota Credit — Our Multiplier ($)",
        "Quota Credit — Full TCV ($)"]
for c, h in enumerate(hdrs, 1):
    col_header(ws_ae, 4, c, h)

term_rows = [
    (1, "1-year", "Inputs!B24"),   # (term_years, label, mult_cell)
    (2, "2-year", "Inputs!B25"),
    (3, "3-year", "Inputs!B26"),
    (5, "5-year", "Inputs!B27"),
]

S1_START = 5
for i, (yrs, lbl, mult_cell) in enumerate(term_rows):
    r = S1_START + i
    label_cell(ws_ae, r, 1, lbl)
    link_cell(ws_ae, r, 2, "=Inputs!B5", FMT_CURRENCY)               # ACV
    formula_cell(ws_ae, r, 3, f"=B{r}*{yrs}", FMT_CURRENCY)           # TCV
    formula_cell(ws_ae, r, 4, f"=B{r}*1", FMT_CURRENCY)               # No mult
    formula_cell(ws_ae, r, 5, f"=B{r}*{mult_cell}", FMT_CURRENCY)     # Our mult
    formula_cell(ws_ae, r, 6, f"=C{r}", FMT_CURRENCY)                 # Full TCV

apply_border_range(ws_ae, 4, 8, 1, 6)

# ── SECTION 2: Bookings Quota Attainment ─────────────────────────────────────
section_header(ws_ae, 10, 1, "SECTION 2 — Bookings Quota Attainment by Approach (% of AE Bookings Quota)", merge_end_col=7)

hdrs2 = ["Term", "AE Bookings Quota ($)",
         "Attainment — No Multiplier",
         "Attainment — Our Multiplier",
         "Attainment — Full TCV"]
for c, h in enumerate(hdrs2, 1):
    col_header(ws_ae, 11, c, h)

S2_START = 12
AE_BK_QUOTA_FORMULA = "=Inputs!B15*Inputs!B16"

for i, (yrs, lbl, mult_cell) in enumerate(term_rows):
    r = S2_START + i
    s1_r = S1_START + i
    label_cell(ws_ae, r, 1, lbl)
    formula_cell(ws_ae, r, 2, AE_BK_QUOTA_FORMULA, FMT_CURRENCY)
    formula_cell(ws_ae, r, 3, f"=D{s1_r}/B{r}", FMT_PCT)   # No mult attainment
    formula_cell(ws_ae, r, 4, f"=E{s1_r}/B{r}", FMT_PCT)   # Our mult attainment
    formula_cell(ws_ae, r, 5, f"=F{s1_r}/B{r}", FMT_PCT)   # Full TCV attainment

apply_border_range(ws_ae, 11, 15, 1, 5)

# Conditional formatting for Section 2
from openpyxl.formatting.rule import CellIsRule
red_fill = PF("solid", fgColor=RED_LIGHT_HEX)
grn_fill = PF("solid", fgColor=GREEN_LIGHT_HEX)

# Full TCV col (E = col 5) — red if > 100%
ws_ae.conditional_formatting.add(
    f"E{S2_START}:E{S2_START+3}",
    CellIsRule(operator="greaterThan", formula=["1"], fill=red_fill))
# Green if >= 80% across attainment cols C,D,E
for col_letter in ["C", "D", "E"]:
    ws_ae.conditional_formatting.add(
        f"{col_letter}{S2_START}:{col_letter}{S2_START+3}",
        CellIsRule(operator="greaterThanOrEqual", formula=["0.8"], fill=grn_fill))

# ── SECTION 3: Annual cACV by Year ────────────────────────────────────────────
section_header(ws_ae, 17, 1, "SECTION 3 — Annual cACV by Year (based on selected contract term)", merge_end_col=7)

hdrs3 = ["Year", "Consumption Rate", "Annual cACV ($)",
         "expansion_signal_acv ($)", "AE cACV Quota Contribution ($)"]
for c, h in enumerate(hdrs3, 1):
    col_header(ws_ae, 18, c, h)

S3_START = 19
rate_cells = ["Inputs!B7", "Inputs!B8", "Inputs!B9", "Inputs!B10", "Inputs!B11"]

for yr in range(1, 6):
    r = S3_START + yr - 1
    rc = rate_cells[yr - 1]
    # Grey out formula: if year > contract term => "N/A"
    in_term = f"(Inputs!B6>={yr})"
    ws_ae.cell(row=r, column=1, value=f"Year {yr}").font = Font(name=ARIAL, size=9)
    # Consumption rate
    c2 = ws_ae.cell(row=r, column=2,
        value=f'=IF({in_term},{rc},"N/A")')
    c2.font = Font(name=ARIAL, size=9, color=GREEN_HEX)
    c2.number_format = FMT_PCT
    # cACV
    c3 = formula_cell(ws_ae, r, 3,
        f'=IF({in_term},MIN(Inputs!B5*{rc},Inputs!B5),"N/A")', FMT_CURRENCY)
    # expansion_signal_acv
    c4 = formula_cell(ws_ae, r, 4,
        f'=IF({in_term},MAX(Inputs!B5*{rc}-Inputs!B5,0),"N/A")', FMT_CURRENCY)
    # AE quota contribution = cACV * AE_cACV_weight
    c5 = formula_cell(ws_ae, r, 5,
        f'=IF({in_term},MIN(Inputs!B5*{rc},Inputs!B5)*Inputs!B17,"N/A")', FMT_CURRENCY)

apply_border_range(ws_ae, 18, 23, 1, 5)

# Grey out rows beyond term — font grey for N/A rows handled by IF formula
# Apply light grey fill to cells that will show N/A
# (We can't do this dynamically per term without VBA; use conditional formatting)
grey_fill = PF("solid", fgColor="F2F2F2")
for yr in range(1, 6):
    r = S3_START + yr - 1
    in_term_cf = f'Inputs!$B$6<{yr}'
    ws_ae.conditional_formatting.add(
        f"A{r}:E{r}",
        CellIsRule(operator="lessThan", formula=[f"Inputs!$B$6+0.1"],
                   fill=grey_fill))

# ── SECTION 4: Consumption Sensitivity ───────────────────────────────────────
section_header(ws_ae, 25, 1, "SECTION 4 — Consumption Sensitivity (flat rate across all contract years)", merge_end_col=7)

hdrs4 = ["Flat Consumption\nRate",
         "Annual cACV ($)",
         "AE cACV Quota\nContribution ($)",
         "AE Bookings\nAttainment % (Our Mult)",
         "Total AE Attainment %\n(Bookings + cACV Yr1)",
         "Delta vs. 20%\nBaseline"]
for c, h in enumerate(hdrs4, 1):
    col_header(ws_ae, 26, c, h)

S4_START = 27
flat_rates = [0.20, 0.40, 0.60, 0.80, 1.00]

# AE bookings attainment (Our Mult) formula — same for all rows
# We need term multiplier based on Inputs!B6
# =IFS(Inputs!B6=1,Inputs!B24,Inputs!B6=2,...) / (Inputs!B15*Inputs!B16)
bk_att_formula = (
    "=(Inputs!B5*IFS(Inputs!B6=1,Inputs!B24,Inputs!B6=2,Inputs!B25,"
    "Inputs!B6=3,Inputs!B26,Inputs!B6=5,Inputs!B27))"
    "/(Inputs!B15*Inputs!B16)"
)

base_total_att_row = None  # row for 20% baseline total attainment
for i, rate in enumerate(flat_rates):
    r = S4_START + i
    # cACV = MIN(ACV * rate, ACV)
    acv_formula = f"=MIN(Inputs!B5*{rate},Inputs!B5)"
    # quota contribution = cACV * AE_cACV_weight
    contrib_formula = f"=B{r}*Inputs!B17"
    # bookings attainment
    bk_att_formula_r = bk_att_formula
    # cACV attainment = contribution / (AE_quota * AE_cACV_weight)
    cacv_att = f"=C{r}/(Inputs!B15*Inputs!B17)"
    total_att = f"=D{r}+{cacv_att[1:]}"  # bookings + cACV yr1

    ic = ws_ae.cell(row=r, column=1, value=rate)
    ic.font = Font(name=ARIAL, size=9, color=BLUE_HEX)
    ic.fill = PatternFill("solid", fgColor=YELLOW_HEX)
    ic.number_format = FMT_PCT

    formula_cell(ws_ae, r, 2, acv_formula, FMT_CURRENCY)
    formula_cell(ws_ae, r, 3, contrib_formula, FMT_CURRENCY)
    formula_cell(ws_ae, r, 4, bk_att_formula_r, FMT_PCT)
    formula_cell(ws_ae, r, 5, total_att, FMT_PCT)

    if i == 0:
        base_total_att_row = r
        formula_cell(ws_ae, r, 6, '="-"', None)  # baseline — no delta
        ws_ae.cell(row=r, column=6).value = "-"
        ws_ae.cell(row=r, column=6).font = Font(name=ARIAL, size=9)
    else:
        formula_cell(ws_ae, r, 6, f"=E{r}-E{base_total_att_row}", FMT_PCT)

apply_border_range(ws_ae, 26, 31, 1, 6)

# ── SECTION 5: Total AE Attainment Summary ────────────────────────────────────
section_header(ws_ae, 33, 1, "SECTION 5 — Total AE Attainment Summary (using Inputs consumption rates)", merge_end_col=7)

hdrs5 = ["Metric",
         "No Multiplier",
         "Our Multiplier",
         "Full TCV"]
for c, h in enumerate(hdrs5, 1):
    col_header(ws_ae, 34, c, h)

S5_START = 35

# Row labels
s5_labels = [
    "Bookings Quota Credit ($)",
    "AE Bookings Quota ($)",
    "Bookings Attainment %",
    "Year 1 cACV ($)",
    "Year 1 cACV Quota Contribution ($)",
    "Year 1 cACV Quota ($)",
    "Year 1 cACV Attainment %",
    "Combined Year 1 Attainment %",
]

# No Mult: ACV*1.0, Our Mult: ACV*mIFS, Full TCV: ACV*term
no_mult_credit  = "=Inputs!B5*1"
our_mult_credit = ("=Inputs!B5*IFS(Inputs!B6=1,Inputs!B24,Inputs!B6=2,Inputs!B25,"
                   "Inputs!B6=3,Inputs!B26,Inputs!B6=5,Inputs!B27)")
full_tcv_credit = "=Inputs!B5*Inputs!B6"
ae_bk_quota     = "=Inputs!B15*Inputs!B16"
yr1_cacv        = "=MIN(Inputs!B5*Inputs!B7,Inputs!B5)"
yr1_contrib     = f"={yr1_cacv[1:]}*Inputs!B17"
ae_cacv_quota   = "=Inputs!B15*Inputs!B17"

rows_data = [
    (no_mult_credit, our_mult_credit, full_tcv_credit, FMT_CURRENCY),
    (ae_bk_quota, ae_bk_quota, ae_bk_quota, FMT_CURRENCY),
    (f"=B{S5_START}/B{S5_START+1}", f"=C{S5_START}/C{S5_START+1}", f"=D{S5_START}/D{S5_START+1}", FMT_PCT),
    (yr1_cacv, yr1_cacv, yr1_cacv, FMT_CURRENCY),
    (yr1_contrib, yr1_contrib, yr1_contrib, FMT_CURRENCY),
    (ae_cacv_quota, ae_cacv_quota, ae_cacv_quota, FMT_CURRENCY),
    (f"=B{S5_START+4}/B{S5_START+5}", f"=C{S5_START+4}/C{S5_START+5}", f"=D{S5_START+4}/D{S5_START+5}", FMT_PCT),
    (f"=B{S5_START+2}+B{S5_START+6}", f"=C{S5_START+2}+C{S5_START+6}", f"=D{S5_START+2}+D{S5_START+6}", FMT_PCT),
]

for i, (lbl, (b_f, c_f, d_f, fmt)) in enumerate(zip(s5_labels, rows_data)):
    r = S5_START + i
    label_cell(ws_ae, r, 1, lbl)
    formula_cell(ws_ae, r, 2, b_f, fmt)
    formula_cell(ws_ae, r, 3, c_f, fmt)
    formula_cell(ws_ae, r, 4, d_f, fmt)

apply_border_range(ws_ae, 34, S5_START+7, 1, 4)

# Highlight combined attainment row with light blue
for col in range(1, 5):
    ws_ae.cell(row=S5_START+7, column=col).fill = PatternFill("solid", fgColor="DDEBF7")

set_col_widths(ws_ae, [28, 18, 18, 24, 24, 22, 22, 5])


# ════════════════════════════════════════════════════════════════════════════
# TAB 3 — AM MODEL
# ════════════════════════════════════════════════════════════════════════════
ws_am = wb.create_sheet("AM Model")
navy_title(ws_am, 1, 1, "Prisma Cloud cACV Comp Model — AM Model", merge_end_col=7)

note_am = ws_am.cell(row=2, column=1,
    value="Note: AM owns cACV from Year 2 onward. Year 1 cACV credit belongs to AE (ramp period).")
note_am.font = Font(name=ARIAL, size=8, italic=True, color="595959")

# ── SECTION 1: AM cACV by Year ───────────────────────────────────────────────
section_header(ws_am, 4, 1, "SECTION 1 — AM cACV by Year (Years 2–5)", merge_end_col=7)

hdrs_am1 = ["Year", "Consumption Rate", "Annual cACV ($)",
            "expansion_signal_acv ($)",
            "AM cACV Quota\nContribution ($)",
            "Cumulative cACV ($)"]
for c, h in enumerate(hdrs_am1, 1):
    col_header(ws_am, 5, c, h)

S1_AM_START = 6

# cumulative: starts from Year 2 (row S1_AM_START)
for yr in range(2, 6):
    r = S1_AM_START + (yr - 2)
    rc = rate_cells[yr - 1]
    in_term = f"(Inputs!B6>={yr})"
    ws_am.cell(row=r, column=1, value=f"Year {yr}").font = Font(name=ARIAL, size=9)
    c2 = ws_am.cell(row=r, column=2, value=f'=IF({in_term},{rc},"N/A")')
    c2.font = Font(name=ARIAL, size=9, color=GREEN_HEX)
    c2.number_format = FMT_PCT
    formula_cell(ws_am, r, 3,
        f'=IF({in_term},MIN(Inputs!B5*{rc},Inputs!B5),"N/A")', FMT_CURRENCY)
    formula_cell(ws_am, r, 4,
        f'=IF({in_term},MAX(Inputs!B5*{rc}-Inputs!B5,0),"N/A")', FMT_CURRENCY)
    formula_cell(ws_am, r, 5,
        f'=IF({in_term},MIN(Inputs!B5*{rc},Inputs!B5)*Inputs!B20,"N/A")', FMT_CURRENCY)
    if yr == 2:
        formula_cell(ws_am, r, 6,
            f'=IF({in_term},MIN(Inputs!B5*{rc},Inputs!B5),"N/A")', FMT_CURRENCY)
    else:
        prev_r = r - 1
        formula_cell(ws_am, r, 6,
            f'=IF({in_term},IF(ISNUMBER(F{prev_r}),F{prev_r},0)+MIN(Inputs!B5*{rc},Inputs!B5),"N/A")',
            FMT_CURRENCY)

apply_border_range(ws_am, 5, S1_AM_START+3, 1, 6)

# ── SECTION 2: AM Bookings — Renewal ─────────────────────────────────────────
section_header(ws_am, 12, 1, "SECTION 2 — AM Bookings: Renewal at End of Contract", merge_end_col=7)

hdrs_am2 = ["Metric", "Value"]
for c, h in enumerate(hdrs_am2, 1):
    col_header(ws_am, 13, c, h)

S2_AM_START = 14
am_renewal_data = [
    ("Renewal ACV ($)",              "=Inputs!B5",                              FMT_CURRENCY, True),
    ("Renewal term (years)",         1,                                         "0",          False),
    ("Renewal multiplier",           "=Inputs!B24",                             "0.00x",      True),
    ("AM bookings quota contribution ($)",
                                    f"=B{S2_AM_START}*B{S2_AM_START+1}*Inputs!B19",
                                                                               FMT_CURRENCY, False),
    ("AM total bookings quota ($)",  "=Inputs!B18*Inputs!B19",                  FMT_CURRENCY, True),
    ("AM bookings attainment %",     f"=B{S2_AM_START+3}/B{S2_AM_START+4}",    FMT_PCT,      False),
]

for i, (lbl, val, fmt, is_link) in enumerate(am_renewal_data):
    r = S2_AM_START + i
    label_cell(ws_am, r, 1, lbl)
    if is_link and isinstance(val, str) and val.startswith("="):
        link_cell(ws_am, r, 2, val, fmt)
    elif isinstance(val, str) and val.startswith("="):
        formula_cell(ws_am, r, 2, val, fmt)
    else:
        input_cell(ws_am, r, 2, val, fmt)

apply_border_range(ws_am, 13, S2_AM_START+5, 1, 2)

# ── SECTION 3: AM Annual Attainment Summary ───────────────────────────────────
section_header(ws_am, 22, 1, "SECTION 3 — AM Annual Attainment Summary", merge_end_col=7)

hdrs_am3 = ["Year",
            "Annual cACV Quota\nContribution ($)",
            "AM cACV Quota ($)",
            "cACV Attainment %\n(This Year)",
            "Cumulative cACV\nAttainment %",
            "Bookings Attainment\n(renewal year only) %",
            "Combined Attainment %"]
for c, h in enumerate(hdrs_am3, 1):
    col_header(ws_am, 23, c, h)

S3_AM_START = 24
AM_CACV_QUOTA = "=Inputs!B18*Inputs!B20"
AM_BK_QUOTA   = "=Inputs!B18*Inputs!B19"
AM_BK_ATT     = f"=B{S2_AM_START+3}/({AM_BK_QUOTA[1:]})"

# renewal year = contract term
for yr in range(2, 6):
    r = S3_AM_START + (yr - 2)
    s1_r = S1_AM_START + (yr - 2)
    in_term = f"(Inputs!B6>={yr})"
    ws_am.cell(row=r, column=1, value=f"Year {yr}").font = Font(name=ARIAL, size=9)
    # Annual cACV quota contribution (link from S1)
    c2 = ws_am.cell(row=r, column=2,
        value=f'=IF({in_term},E{s1_r},"N/A")')
    c2.font = Font(name=ARIAL, size=9, color=GREEN_HEX)
    c2.number_format = FMT_CURRENCY
    # AM cACV quota
    c3 = formula_cell(ws_am, r, 3,
        f'=IF({in_term},{AM_CACV_QUOTA[1:]},"N/A")', FMT_CURRENCY)
    # cACV attainment this year
    c4 = formula_cell(ws_am, r, 4,
        f'=IF({in_term},B{r}/C{r},"N/A")', FMT_PCT)
    # cumulative attainment = cumulative cACV contrib / cACV quota
    # cumulative cACV contrib = F{s1_r} * AM_cACV_weight  (cumulative cACV * weight)
    c5 = formula_cell(ws_am, r, 5,
        f'=IF({in_term},(IF(ISNUMBER(F{s1_r}),F{s1_r},0)*Inputs!B20)/({AM_CACV_QUOTA[1:]}),"N/A")',
        FMT_PCT)
    # Bookings attainment — only at renewal year (= contract term)
    c6 = formula_cell(ws_am, r, 6,
        f'=IF(Inputs!B6={yr},{AM_BK_ATT[1:]},"N/A")', FMT_PCT)
    # Combined attainment
    c7 = formula_cell(ws_am, r, 7,
        f'=IF({in_term},D{r}+IF(ISNUMBER(F{r}),F{r},0),"N/A")', FMT_PCT)

apply_border_range(ws_am, 23, S3_AM_START+3, 1, 7)

# Conditional formatting: green if combined attainment >= 80%
grn_fill2 = PF("solid", fgColor=GREEN_LIGHT_HEX)
ws_am.conditional_formatting.add(
    f"G{S3_AM_START}:G{S3_AM_START+3}",
    CellIsRule(operator="greaterThanOrEqual", formula=["0.8"], fill=grn_fill2))

set_col_widths(ws_am, [28, 18, 18, 22, 22, 24, 22, 5])

# ── Apply font to all cells globally ─────────────────────────────────────────
for ws in [ws_inp, ws_ae, ws_am]:
    for row in ws.iter_rows():
        for cell in row:
            if cell.font:
                old = cell.font
                if old.name != ARIAL:
                    cell.font = Font(
                        name=ARIAL, bold=old.bold, italic=old.italic,
                        size=old.size, color=old.color,
                        underline=old.underline)

# ── Save ─────────────────────────────────────────────────────────────────────
out = "/Users/jantai/Documents/Claude/Projects/gtm/comp_model.xlsx"
wb.save(out)
print(f"Saved: {out}")
